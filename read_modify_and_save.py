from openpyxl import load_workbook


workbook = load_workbook("created_excel_file.xlsx")
sheet = workbook.active
sheet["A1"] = "Changed_text"
workbook.create_sheet("New_sheet")
sheet = workbook.get_sheet_by_name('New_sheet')
sheet["B1"] = "text in new sheet"
workbook.save("created_excel_file_result.xlsx")
